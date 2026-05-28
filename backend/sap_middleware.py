"""
LogiStep SAP Middleware
-----------------------
Mobilden gelen teslimat onayini SAP ERP sistemine (RFC/BAPI) ileten ve
sonuclari Excel uyum raporuna yazan otomasyon katmanidir.

NOT: pyrfc kutuphanesi SAP NetWeaver RFC SDK gerektirir. SDK kurulu degilse
modul otomatik olarak SIMULASYON modunda calisir; boylece demo/portfolyo
amaciyla gercek bir SAP sistemi olmadan da uctan uca calistirilabilir.
"""
import os
import datetime
import logging

from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

logger = logging.getLogger("logistep.sap")

# pyrfc opsiyonel: kurulu degilse simulasyon moduna dus
try:
    from pyrfc import Connection, RFCError

    PYRFC_AVAILABLE = True
except ImportError:  # pragma: no cover
    PYRFC_AVAILABLE = False

    class RFCError(Exception):
        """pyrfc kurulu olmadiginda yer tutucu istisna."""

        pass


load_dotenv()

REPORT_FILE = "Delivery_Compliance_Report.xlsx"


class SAPDeliveryManager:
    def __init__(self, simulate: bool = None):
        self.conn_params = {
            "user": os.getenv("SAP_USER"),
            "passwd": os.getenv("SAP_PASSWORD"),
            "ashost": os.getenv("SAP_HOST"),
            "sysnr": os.getenv("SAP_SYSNR"),
            "client": os.getenv("SAP_CLIENT"),
        }

        # pyrfc yoksa veya baglanti bilgileri eksikse simulasyona dus
        if simulate is None:
            simulate = (not PYRFC_AVAILABLE) or (not all(self.conn_params.values()))
        self.simulate = simulate

        if self.simulate:
            logger.warning(
                "SAP middleware SIMULASYON modunda calisiyor "
                "(gercek SAP baglantisi yok)."
            )

        self._init_report()

    def _init_report(self):
        """Rapor dosyasi yoksa baslikla olustur, varsa mevcut olana ekle."""
        if os.path.exists(REPORT_FILE):
            self.workbook = load_workbook(REPORT_FILE)
            self.sheet = self.workbook.active
        else:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = "Compliance"
            self.sheet.append(["Zeitstempel", "Liefer-ID", "Status", "Nachricht"])
            self.workbook.save(REPORT_FILE)

    def connect(self):
        if self.simulate:
            return None
        try:
            return Connection(**self.conn_params)
        except RFCError as e:
            logger.error("SAP Baglanti Hatasi: %s", e)
            return None

    def update_delivery(self, delivery_id: str, status_code: str) -> dict:
        """Mobil cihazdan gelen onayla BAPI'yi tetikler. Sonucu dict olarak doner."""
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # --- SIMULASYON MODU (SAP'siz calistirma) ---
        if self.simulate:
            msg = f"[SIMULATION] Lieferung {delivery_id} -> {status_code}"
            self._log_to_excel(timestamp, delivery_id, "SIMULIERT", msg)
            logger.info(msg)
            return {"success": True, "status": status_code, "message": msg}

        # --- GERCEK SAP BAGLANTISI ---
        conn = self.connect()
        if not conn:
            msg = "Keine Verbindung zum SAP-System"
            self._log_to_excel(timestamp, delivery_id, "FEHLER", msg)
            return {"success": False, "status": None, "message": msg}

        try:
            conn.call(
                "BAPI_OUTB_DELIVERY_CREATE_STO",
                DELIVERY_ID=delivery_id,
                STATUS=status_code,
            )
            conn.call("BAPI_TRANSACTION_COMMIT")
            msg = "Lieferung erfolgreich aktualisiert"
            self._log_to_excel(timestamp, delivery_id, "ERFOLG", msg)
            logger.info("[%s] Teslimat %s SAP sistemine islendi.", timestamp, delivery_id)
            return {"success": True, "status": status_code, "message": msg}
        except RFCError as e:
            conn.call("BAPI_TRANSACTION_ROLLBACK")
            self._log_to_excel(timestamp, delivery_id, "FEHLER", str(e))
            return {"success": False, "status": None, "message": str(e)}
        finally:
            conn.close()

    def _log_to_excel(self, timestamp, delivery_id, status, message):
        """Islem sonuclarini Excel raporuna yazar."""
        self.sheet.append([timestamp, delivery_id, status, message])
        self.workbook.save(REPORT_FILE)


# --- Test Calistirmasi (Simulasyon) ---
if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s"
    )
    print("LogiStep SAP Middleware Baslatiliyor...")
    manager = SAPDeliveryManager()
    result = manager.update_delivery("80012345", "DELIVERED")
    print("Sonuc:", result)
